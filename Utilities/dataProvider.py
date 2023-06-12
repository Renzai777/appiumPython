import openpyxl


# def get_data(sheetName):
#     workbook = openpyxl.load_workbook("../Excel/testdata.xlsx")
#     sheet = workbook[sheetName]
#     totalrows = sheet.max_row
#     totalcols = sheet.max_column
#
#     print("total cols are ",str(totalcols))
#     print("total rows are ", str(totalrows))
#     mainList = []
#
#     for i in range(2,totalrows+1):
#         dataList = []
#         for j in range(1, totalcols+1):
#             data = sheet.cell(row=i,column=j).value
#             dataList.insert(j,data)
#         mainList.insert(i,dataList)
#         print(mainList)
#     return mainList




def get_data(sheetName):
    workbook = openpyxl.load_workbook("../Excel/testdata.xlsx")
    sheet = workbook[sheetName]
    totalrows = sheet.max_row
    totalcols = sheet.max_column
    print("total cols are ",str(totalcols))
    print("total rows are ", str(totalrows))
    mainList = []

    for i in range(2,totalrows+1):
        dataList = []
        for j in range(1, totalcols+1):
            data = sheet.cell(row=i,column=j).value
            dataList.insert(j,data)
        mainList.insert(i,dataList)
        print(mainList)
    return mainList









#---------tuple
# def get_data(sheetName):
#     workbook = openpyxl.load_workbook("../Excel/testdata.xlsx")
#     sheet = workbook[sheetName]
#     totalrows = sheet.max_row
#     totalcols = sheet.max_column
#
#     print("total cols are ", str(totalcols))
#     print("total rows are ", str(totalrows))
#     mainList = []
#
#     for i in range(2, totalrows + 1):
#         dataList = []
#         for j in range(1, totalcols + 1):
#             data = sheet.cell(row=i, column=j).value
#             dataList.append(data)
#         mainList.append(tuple(dataList))
#         print(mainList)
#     return mainList


#

# def get_data(sheetName):
#     workbook = openpyxl.load_workbook("../Excel/testdata.xlsx")
#     sheet = workbook[sheetName]
#     totalrows = sheet.max_row
#     totalcols = sheet.max_column
#
#     print("total cols are ", str(totalcols))
#     print("total rows are ", str(totalrows))
#     mainList = []
#
#     for i in range(2, totalrows + 1):
#         row_data = []
#         is_empty_row = True
#         for j in range(1, totalcols + 1):
#             data = sheet.cell(row=i, column=j).value
#             row_data.append(data)
#             if data:
#                 is_empty_row = False
#         if not is_empty_row:
#             mainList.append(row_data)
#         print(mainList)
#
#     return mainList





# def get_data(sheetName):
#     workbook = openpyxl.load_workbook("../Excel/testdata.xlsx")
#     sheet = workbook[sheetName]
#     totalrows = sheet.max_row
#     totalcols = sheet.max_column
#
#     print("total cols are ", str(totalcols))
#     print("total rows are ", str(totalrows))
#     mainList = []
#
#     for i in range(2, totalrows + 1):
#         dataList = []
#         for j in range(1, totalcols + 1):
#             data = sheet.cell(row=i, column=j).value
#             dataList.append(data)
#         mainList.append(tuple(dataList))
#         print(mainList)
#     return mainList



get_data("loginTest")
# print(type(get_data("loginTest")))
